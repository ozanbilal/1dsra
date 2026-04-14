from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

GRAVITY_M_S2 = 9.81


@dataclass(slots=True)
class DeepsoilExcelBundle:
    workbook_path: Path
    output_dir: Path
    case_kind: str
    sheet_map: dict[str, str]
    column_map: dict[str, dict[str, int]]
    available_artifacts: list[str]
    warnings: list[str]
    surface_csv: Path | None = None
    input_motion_csv: Path | None = None
    psa_surface_csv: Path | None = None
    psa_input_csv: Path | None = None
    profile_csv: Path | None = None
    mobilized_strength_csv: Path | None = None
    hysteresis_csv: Path | None = None
    meta_json: Path | None = None


def _norm(value: object) -> str:
    return str(value or "").strip().casefold()


def _as_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _write_csv(path: Path, headers: list[str], rows: list[list[float | int]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        writer.writerows(rows)


def _find_sheet_name(sheet_names: list[str], target: str) -> str | None:
    target_norm = _norm(target)
    for name in sheet_names:
        if _norm(name) == target_norm:
            return name
    return None


def _header_index_map(header_row: tuple[object, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, value in enumerate(header_row):
        key = _norm(value)
        if key:
            mapping[key] = idx
    return mapping


def _parse_layer1_sheet(ws: Any, output_dir: Path) -> tuple[dict[str, Path], dict[str, int], list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Layer 1 sheet is empty.")
    header = rows[0]
    col_map = _header_index_map(header)
    warnings: list[str] = []

    time_idx = col_map.get("time (s)")
    acc_idx = col_map.get("acceleration (g)")
    strain_idx = col_map.get("strain (%)")
    stress_idx = col_map.get("shear stress(kpa)")
    period_idx = col_map.get("period (sec)")
    psa_idx = col_map.get("psa (g)")

    if time_idx is None or acc_idx is None:
        raise ValueError("Layer 1 sheet must contain Time (s) and Acceleration (g).")

    surface_rows: list[list[float]] = []
    psa_rows: list[list[float]] = []
    hysteresis_rows: list[list[float]] = []

    for row in rows[1:]:
        time_val = _as_float(row[time_idx]) if time_idx < len(row) else None
        acc_val = _as_float(row[acc_idx]) if acc_idx < len(row) else None
        if time_val is not None and acc_val is not None:
            surface_rows.append([time_val, acc_val * GRAVITY_M_S2])

        period_val = _as_float(row[period_idx]) if period_idx is not None and period_idx < len(row) else None
        psa_val = _as_float(row[psa_idx]) if psa_idx is not None and psa_idx < len(row) else None
        if period_val is not None and psa_val is not None and period_val > 0.0:
            psa_rows.append([period_val, psa_val * GRAVITY_M_S2])

        strain_val = _as_float(row[strain_idx]) if strain_idx is not None and strain_idx < len(row) else None
        stress_val = _as_float(row[stress_idx]) if stress_idx is not None and stress_idx < len(row) else None
        if strain_val is not None and stress_val is not None:
            # DeepSoil workbook exports shear stress directly in kPa.
            hysteresis_rows.append([strain_val / 100.0, stress_val])

    artifacts: dict[str, Path] = {}
    if surface_rows:
        artifacts["surface_csv"] = output_dir / "surface.csv"
        _write_csv(artifacts["surface_csv"], ["time_s", "acc_m_s2"], surface_rows)
    else:
        warnings.append("Layer 1 sheet did not yield surface time-history rows.")
    if psa_rows:
        artifacts["psa_surface_csv"] = output_dir / "psa_surface.csv"
        _write_csv(artifacts["psa_surface_csv"], ["period_s", "psa_m_s2"], psa_rows)
    else:
        warnings.append("Layer 1 sheet did not yield surface PSA rows.")
    if hysteresis_rows:
        artifacts["hysteresis_csv"] = output_dir / "hysteresis_layer1.csv"
        _write_csv(artifacts["hysteresis_csv"], ["strain", "stress"], hysteresis_rows)
    else:
        warnings.append("Layer 1 sheet did not yield hysteresis rows.")
    return artifacts, {k: v + 1 for k, v in col_map.items()}, warnings


def _parse_input_motion_sheet(ws: Any, output_dir: Path) -> tuple[dict[str, Path], dict[str, int], list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Input Motion sheet is empty.")
    header = rows[1]
    col_map = _header_index_map(header)
    warnings: list[str] = []

    time_idx = col_map.get("time (s)")
    acc_idx = col_map.get("acceleration (g)")
    period_idx = col_map.get("period (sec)")
    psa_idx = col_map.get("psa (g)")

    if time_idx is None or acc_idx is None:
        raise ValueError("Input Motion sheet must contain Time (s) and Acceleration (g).")

    input_rows: list[list[float]] = []
    psa_rows: list[list[float]] = []

    for row in rows[2:]:
        time_val = _as_float(row[time_idx]) if time_idx < len(row) else None
        acc_val = _as_float(row[acc_idx]) if acc_idx < len(row) else None
        if time_val is not None and acc_val is not None:
            input_rows.append([time_val, acc_val * GRAVITY_M_S2])

        period_val = _as_float(row[period_idx]) if period_idx is not None and period_idx < len(row) else None
        psa_val = _as_float(row[psa_idx]) if psa_idx is not None and psa_idx < len(row) else None
        if period_val is not None and psa_val is not None and period_val > 0.0:
            psa_rows.append([period_val, psa_val * GRAVITY_M_S2])

    artifacts: dict[str, Path] = {}
    if input_rows:
        artifacts["input_motion_csv"] = output_dir / "input_motion.csv"
        _write_csv(artifacts["input_motion_csv"], ["time_s", "acc_m_s2"], input_rows)
    else:
        warnings.append("Input Motion sheet did not yield input history rows.")
    if psa_rows:
        artifacts["psa_input_csv"] = output_dir / "psa_input.csv"
        _write_csv(artifacts["psa_input_csv"], ["period_s", "psa_m_s2"], psa_rows)
    else:
        warnings.append("Input Motion sheet did not yield input PSA rows.")
    return artifacts, {k: v + 1 for k, v in col_map.items()}, warnings


def _profile_field_from_title(title: str) -> str | None:
    normalized = _norm(title)
    if normalized.startswith("effective stress"):
        return "effective_stress_kpa"
    if normalized.startswith("pga"):
        return "pga_g"
    if "maximum displacement" in normalized:
        return "max_displacement_m"
    if "max. strain" in normalized or "max strain" in normalized:
        return "max_strain_pct"
    if "max. stress ratio" in normalized or "max stress ratio" in normalized:
        return "max_stress_ratio"
    return None


def _parse_profile_sheet(ws: Any, output_dir: Path) -> tuple[dict[str, Path], dict[str, int], list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        raise ValueError("Profile sheet is empty.")
    title_row = rows[0]
    header_row = rows[1]
    warnings: list[str] = []

    block_columns: dict[str, int] = {}
    records: dict[float, dict[str, float]] = {}

    for idx, title in enumerate(title_row):
        field_name = _profile_field_from_title(str(title or ""))
        if field_name is None:
            continue
        block_columns[field_name] = idx + 1
        for row in rows[2:]:
            depth_val = _as_float(row[idx]) if idx < len(row) else None
            value_val = _as_float(row[idx + 1]) if (idx + 1) < len(row) else None
            if depth_val is None or value_val is None:
                continue
            key = round(depth_val, 8)
            rec = records.setdefault(key, {"depth_m": depth_val})
            rec[field_name] = value_val

    if not records:
        warnings.append("Profile sheet did not yield any profile rows.")
        return {}, block_columns, warnings

    ordered_rows = [records[key] for key in sorted(records)]
    headers = [
        "depth_m",
        "effective_stress_kpa",
        "pga_g",
        "max_displacement_m",
        "max_strain_pct",
        "max_stress_ratio",
    ]
    out_rows: list[list[float]] = []
    for record in ordered_rows:
        out_rows.append([float(record.get(col, float("nan"))) for col in headers])
    artifact = output_dir / "profile.csv"
    _write_csv(artifact, headers, out_rows)
    return {"profile_csv": artifact}, block_columns, warnings


def _parse_mobilized_sheet(ws: Any, output_dir: Path) -> tuple[dict[str, Path], dict[str, int], list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Mobilized Shear Stress sheet is empty.")
    header = rows[0]
    col_map = _header_index_map(header)
    warnings: list[str] = []

    layer_idx = col_map.get("layer")
    strength_idx = col_map.get("shear strength (kpa)")
    friction_idx = col_map.get("friction angle (deg)")
    if layer_idx is None or strength_idx is None or friction_idx is None:
        raise ValueError(
            "Mobilized Shear Stress sheet must contain Layer, Shear Strength (kPa), and Friction Angle (deg)."
        )

    mobilized_rows: list[list[float | int]] = []
    for row in rows[1:]:
        layer = _as_float(row[layer_idx]) if layer_idx < len(row) else None
        strength = _as_float(row[strength_idx]) if strength_idx < len(row) else None
        friction = _as_float(row[friction_idx]) if friction_idx < len(row) else None
        if layer is None or strength is None or friction is None:
            continue
        mobilized_rows.append([int(layer), strength, friction])

    artifacts: dict[str, Path] = {}
    if mobilized_rows:
        artifacts["mobilized_strength_csv"] = output_dir / "mobilized_strength.csv"
        _write_csv(
            artifacts["mobilized_strength_csv"],
            ["layer", "shear_strength_kpa", "friction_angle_deg"],
            mobilized_rows,
        )
    else:
        warnings.append("Mobilized Shear Stress sheet did not yield mobilized rows.")
    return artifacts, {k: v + 1 for k, v in col_map.items()}, warnings


def import_deepsoil_excel_bundle(
    workbook_path: str | Path,
    output_dir: str | Path,
    *,
    case_kind: str | None = None,
) -> DeepsoilExcelBundle:
    workbook_file = Path(workbook_path)
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    inferred_kind = "secondary_el" if "-el" in workbook_file.stem.casefold() else "primary_gqh"
    bundle = DeepsoilExcelBundle(
        workbook_path=workbook_file,
        output_dir=out_dir,
        case_kind=case_kind or inferred_kind,
        sheet_map={},
        column_map={},
        available_artifacts=[],
        warnings=[],
    )

    workbook = load_workbook(workbook_file, data_only=True, read_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        required = {
            "layer_1": "Layer 1",
            "input_motion": "Input Motion",
            "profile": "Profile",
            "mobilized": "Mobilized Shear Stress",
        }
        for key, expected in required.items():
            actual = _find_sheet_name(sheet_names, expected)
            if actual is None:
                bundle.warnings.append(f"Workbook is missing expected sheet '{expected}'.")
            else:
                bundle.sheet_map[key] = actual

        if "layer_1" in bundle.sheet_map:
            artifacts, column_map, warnings = _parse_layer1_sheet(
                workbook[bundle.sheet_map["layer_1"]],
                out_dir,
            )
            bundle.column_map["layer_1"] = column_map
            bundle.warnings.extend(warnings)
            bundle.surface_csv = artifacts.get("surface_csv")
            bundle.psa_surface_csv = artifacts.get("psa_surface_csv")
            bundle.hysteresis_csv = artifacts.get("hysteresis_csv")

        if "input_motion" in bundle.sheet_map:
            artifacts, column_map, warnings = _parse_input_motion_sheet(
                workbook[bundle.sheet_map["input_motion"]],
                out_dir,
            )
            bundle.column_map["input_motion"] = column_map
            bundle.warnings.extend(warnings)
            bundle.input_motion_csv = artifacts.get("input_motion_csv")
            bundle.psa_input_csv = artifacts.get("psa_input_csv")

        if "profile" in bundle.sheet_map:
            artifacts, column_map, warnings = _parse_profile_sheet(
                workbook[bundle.sheet_map["profile"]],
                out_dir,
            )
            bundle.column_map["profile"] = column_map
            bundle.warnings.extend(warnings)
            bundle.profile_csv = artifacts.get("profile_csv")

        if "mobilized" in bundle.sheet_map:
            artifacts, column_map, warnings = _parse_mobilized_sheet(
                workbook[bundle.sheet_map["mobilized"]],
                out_dir,
            )
            bundle.column_map["mobilized"] = column_map
            bundle.warnings.extend(warnings)
            bundle.mobilized_strength_csv = artifacts.get("mobilized_strength_csv")
    finally:
        workbook.close()

    available = {
        "surface.csv": bundle.surface_csv,
        "input_motion.csv": bundle.input_motion_csv,
        "psa_surface.csv": bundle.psa_surface_csv,
        "psa_input.csv": bundle.psa_input_csv,
        "profile.csv": bundle.profile_csv,
        "mobilized_strength.csv": bundle.mobilized_strength_csv,
        "hysteresis_layer1.csv": bundle.hysteresis_csv,
    }
    bundle.available_artifacts = [name for name, path in available.items() if path is not None]

    meta_path = out_dir / "meta.json"
    meta_path.write_text(
        json.dumps(
            {
                "source_workbook": str(workbook_file),
                "sheet_map": bundle.sheet_map,
                "column_map": bundle.column_map,
                "available_artifacts": bundle.available_artifacts,
                "warnings": bundle.warnings,
                "case_kind": bundle.case_kind,
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    bundle.meta_json = meta_path
    return bundle
