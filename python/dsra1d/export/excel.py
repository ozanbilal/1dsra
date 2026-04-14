"""Multi-sheet Excel (.xlsx) export for GeoWave run results.

Produces a DEEPSOIL-equivalent workbook with:
  Sheet 1 - Summary: project metadata, PGA, solver info
  Sheet 2 - Time History: time, surface_acc, input_acc
  Sheet 3 - Spectral: periods, PSA_surface, PSA_input
  Sheet 4 - Profile: layer depth, Vs, unit weight, max strain, damping
  Sheet 5 - Stress-Strain: per-layer hysteretic loop data
  Sheet 6 - EQL Convergence: iteration history (EQL only)
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _header_style() -> tuple[Font, PatternFill, Alignment]:
    """Return consistent header styling."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export: pip install openpyxl")
    font = Font(bold=True, size=11)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font_white = Font(bold=True, size=11, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")
    return font_white, fill, alignment


def _auto_width(ws: Any) -> None:
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def _write_header_row(ws: Any, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    font, fill, alignment = _header_style()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment


def _write_summary_sheet(
    ws: Any,
    run_meta: dict[str, Any],
    config_snapshot: dict[str, Any],
    pga: float,
    pro_metrics: dict[str, Any] | None = None,
) -> None:
    """Write project summary sheet."""
    ws.title = "Summary"

    info_rows = [
        ("Project Name", config_snapshot.get("project_name", "GeoWave Analysis")),
        ("Run ID", run_meta.get("run_id", "N/A")),
        ("Solver Backend", run_meta.get("backend", "N/A")),
        ("Boundary Condition", config_snapshot.get("analysis", {}).get("boundary_condition", "N/A")),
        ("Time Step (s)", run_meta.get("delta_t_s", run_meta.get("dt_s", "N/A"))),
        ("PGA (m/s2)", f"{pga:.6f}"),
        ("PGA (g)", f"{pga / 9.81:.6f}"),
        ("Number of Layers", len(config_snapshot.get("profile", {}).get("layers", []))),
        ("Motion File", config_snapshot.get("motion", {}).get("file", "N/A")),
        ("Damping Mode", config_snapshot.get("analysis", {}).get("damping_mode", "N/A")),
    ]

    # Pro metrics
    if pro_metrics:
        if pro_metrics.get("site_period_s") is not None:
            info_rows.append(("Site Period T₀ (s)", f"{pro_metrics['site_period_s']:.4f}"))
        if pro_metrics.get("vs_avg_m_s") is not None:
            info_rows.append(("Average Vs (m/s)", f"{pro_metrics['vs_avg_m_s']:.1f}"))
        if pro_metrics.get("kappa") is not None:
            info_rows.append(("Kappa κ (s)", f"{pro_metrics['kappa']:.6f}"))
        if pro_metrics.get("kappa_r2") is not None:
            info_rows.append(("Kappa R²", f"{pro_metrics['kappa_r2']:.4f}"))

    _write_header_row(ws, 1, ["Parameter", "Value"])
    for row_idx, (param, value) in enumerate(info_rows, 2):
        ws.cell(row=row_idx, column=1, value=param)
        ws.cell(row=row_idx, column=2, value=str(value))

    _auto_width(ws)


def _write_time_history_sheet(
    ws: Any,
    time: np.ndarray,
    surface_acc: np.ndarray,
    input_acc: np.ndarray | None = None,
) -> None:
    """Write time history data."""
    ws.title = "Time History"

    headers = ["Time (s)", "Surface Acc (m/s2)", "Surface Acc (g)"]
    if input_acc is not None:
        headers.extend(["Input Acc (m/s2)", "Input Acc (g)"])

    _write_header_row(ws, 1, headers)

    n = min(len(time), len(surface_acc))
    for i in range(n):
        ws.cell(row=i + 2, column=1, value=float(time[i]))
        ws.cell(row=i + 2, column=2, value=float(surface_acc[i]))
        ws.cell(row=i + 2, column=3, value=float(surface_acc[i] / 9.81))
        if input_acc is not None and i < len(input_acc):
            ws.cell(row=i + 2, column=4, value=float(input_acc[i]))
            ws.cell(row=i + 2, column=5, value=float(input_acc[i] / 9.81))


def _write_spectral_sheet(
    ws: Any,
    periods: np.ndarray,
    psa_surface: np.ndarray,
    psa_input: np.ndarray | None = None,
    include_psv_psd: bool = True,
) -> None:
    """Write response spectra data with optional PSV/PSD columns."""
    ws.title = "Spectral"

    headers = ["Period (s)", "Frequency (Hz)", "PSA Surface (m/s2)", "PSA Surface (g)"]
    if include_psv_psd:
        headers.extend(["PSV Surface (m/s)", "PSD Surface (m)"])
    if psa_input is not None:
        headers.extend(["PSA Input (m/s2)", "PSA Input (g)", "Amplification"])

    _write_header_row(ws, 1, headers)

    two_pi = 2.0 * 3.141592653589793
    four_pi2 = 4.0 * 3.141592653589793 ** 2
    for i in range(len(periods)):
        row = i + 2
        p = float(periods[i])
        ws.cell(row=row, column=1, value=p)
        ws.cell(row=row, column=2, value=1.0 / p if p > 0 else 0.0)
        sa = float(psa_surface[i]) if i < len(psa_surface) else 0.0
        ws.cell(row=row, column=3, value=sa)
        ws.cell(row=row, column=4, value=sa / 9.81)
        col = 5
        if include_psv_psd:
            ws.cell(row=row, column=col, value=sa * p / two_pi)       # PSV
            ws.cell(row=row, column=col + 1, value=sa * p ** 2 / four_pi2)  # PSD
            col += 2
        if psa_input is not None and i < len(psa_input):
            sa_in = float(psa_input[i])
            ws.cell(row=row, column=col, value=sa_in)
            ws.cell(row=row, column=col + 1, value=sa_in / 9.81)
            amp = sa / sa_in if sa_in > 1e-12 else 0.0
            ws.cell(row=row, column=col + 2, value=amp)


def _write_profile_sheet(
    ws: Any,
    layers: list[dict[str, Any]],
    strains: np.ndarray | None = None,
) -> None:
    """Write layer profile summary."""
    ws.title = "Profile"

    headers = [
        "Layer", "Depth Top (m)", "Depth Bottom (m)", "Thickness (m)",
        "Vs (m/s)", "Unit Weight (kN/m3)", "Material",
        "Max Strain", "Damping Min", "Damping Max",
    ]
    _write_header_row(ws, 1, headers)

    depth = 0.0
    for i, layer in enumerate(layers):
        row = i + 2
        thickness = layer.get("thickness", 0.0)
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=depth)
        ws.cell(row=row, column=3, value=depth + thickness)
        ws.cell(row=row, column=4, value=thickness)
        ws.cell(row=row, column=5, value=layer.get("vs", 0.0))
        ws.cell(row=row, column=6, value=layer.get("unit_weight", 0.0))
        ws.cell(row=row, column=7, value=layer.get("material", "N/A"))
        mp = layer.get("material_params", {})
        strain_val = float(strains[i]) if strains is not None and i < len(strains) else 0.0
        ws.cell(row=row, column=8, value=strain_val)
        ws.cell(row=row, column=9, value=mp.get("damping_min", 0.0))
        ws.cell(row=row, column=10, value=mp.get("damping_max", 0.0))
        depth += thickness

    _auto_width(ws)


def _write_eql_convergence_sheet(
    ws: Any,
    eql_summary: dict[str, Any],
) -> None:
    """Write EQL convergence history."""
    ws.title = "EQL Convergence"

    headers = ["Iteration", "Max Vs Change (%)", "Converged"]
    _write_header_row(ws, 1, headers)

    history = eql_summary.get("max_change_history", [])
    converged = eql_summary.get("converged", False)
    for i, change in enumerate(history):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=float(change) * 100.0)
        ws.cell(row=row, column=3, value="Yes" if i == len(history) - 1 and converged else "No")

    _auto_width(ws)


def export_run_to_xlsx(
    run_dir: Path,
    output_path: Path | None = None,
    include_pro_sheets: bool = True,
) -> Path:
    """Export a GeoWave run to multi-sheet Excel workbook.

    Parameters
    ----------
    run_dir : Path
        Path to the run directory containing results.
    output_path : Path | None
        Output .xlsx path. Defaults to ``run_dir / "results.xlsx"``.

    Returns
    -------
    Path
        Path to the generated .xlsx file.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export: pip install openpyxl")

    from dsra1d.store import load_result

    store = load_result(run_dir)
    out = output_path or (run_dir / "results.xlsx")

    # Load metadata
    meta_path = run_dir / "run_meta.json"
    run_meta: dict[str, Any] = {}
    if meta_path.exists():
        run_meta = json.loads(meta_path.read_text(encoding="utf-8"))

    config_path = run_dir / "config_snapshot.json"
    config_snapshot: dict[str, Any] = {}
    if config_path.exists():
        config_snapshot = json.loads(config_path.read_text(encoding="utf-8"))

    pga = float(np.max(np.abs(store.acc_surface))) if store.acc_surface.size > 0 else 0.0

    # Compute Pro metrics
    pro_metrics: dict[str, Any] = {}
    try:
        prof_layers = config_snapshot.get("profile", {}).get("layers", [])
        if prof_layers:
            total_h = sum(float(la.get("thickness_m", 0)) for la in prof_layers)
            travel_time = sum(
                float(la.get("thickness_m", 0)) / max(float(la.get("vs_m_s", 100)), 1.0)
                for la in prof_layers
            )
            if travel_time > 0 and total_h > 0:
                pro_metrics["vs_avg_m_s"] = total_h / travel_time
                pro_metrics["site_period_s"] = 4.0 * total_h / pro_metrics["vs_avg_m_s"]
    except Exception:
        pass

    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    assert ws_summary is not None
    _write_summary_sheet(ws_summary, run_meta, config_snapshot, pga, pro_metrics=pro_metrics)

    # Sheet 2: Time History
    ws_th = wb.create_sheet()
    input_acc_for_export = store.acc_input if store.acc_input.size > 1 else None
    _write_time_history_sheet(ws_th, store.time, store.acc_surface, input_acc=input_acc_for_export)

    # Sheet 3: Spectral
    from dsra1d.post.spectra import compute_spectra

    periods = np.logspace(np.log10(0.01), np.log10(10.0), 100)
    dt = float(store.dt_s)
    spectra_result = compute_spectra(store.acc_surface, dt, damping=0.05, periods=periods)
    psa_surface = spectra_result.psa
    ws_spec = wb.create_sheet()
    _write_spectral_sheet(ws_spec, periods, psa_surface, include_psv_psd=include_pro_sheets)

    # Sheet 4: Profile
    layers_data = config_snapshot.get("profile", {}).get("layers", [])
    max_strains = store.max_strains if hasattr(store, "max_strains") else None
    ws_prof = wb.create_sheet()
    _write_profile_sheet(ws_prof, layers_data, max_strains)

    # Sheet 5: EQL Convergence (if available)
    eql_path = run_dir / "eql_summary.json"
    if eql_path.exists():
        eql_data = json.loads(eql_path.read_text(encoding="utf-8"))
        ws_eql = wb.create_sheet()
        _write_eql_convergence_sheet(ws_eql, eql_data)

    wb.save(str(out))
    return out
